"""
Filename    : upbitMT.py [/upbit-MT/]
Author      : [Jin2ouS]
Date        : 2025-02-01
Version     : 1.0.0
Description : 업비트 Open API를 활용한 크립토 자동 감시(Monitoring) 및 주문 실행(Trading)
              - upbitMT.py: 단일 실행 파일
              - utils.py: 메시지(Slack/Telegram)
              - .env: 환경설정 (API 키, 알림 채널 등)
              - upbitMT.list.xlsx: 임계값 엑셀 (autoMT.list.KIS.xlsx 형식)
"""

import os
import sys

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
import re
import math
import uuid
import json
import time
import hashlib
import atexit
import signal
import sys
from datetime import datetime
from urllib.parse import urlencode, unquote

import requests
import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv
import jwt

from utils import get_runtime_info, send_message

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
load_dotenv(os.path.join(BASE_DIR, ".env"))

UPBIT_ACCESS_KEY = os.getenv("UPBIT_ACCESS_KEY", "").strip()
UPBIT_SECRET_KEY = os.getenv("UPBIT_SECRET_KEY", "").strip()
MONITOR_FILE = os.getenv("MONITOR_FILE", "upbitMT.list.xlsx").strip()
EXCEL_PATH = os.path.join(BASE_DIR, MONITOR_FILE) if not os.path.isabs(MONITOR_FILE) else MONITOR_FILE

UPBIT_BASE_URL = "https://api.upbit.com/v1"
SCRIPT_NAME = os.path.basename(__file__)

NUM_CANDLES_PRICE = 1
SEND_HOURLY_MSG = False
UPBIT_MIN_ORDER_KRW = 5000

if not UPBIT_ACCESS_KEY or not UPBIT_SECRET_KEY:
    raise ValueError("UPBIT_ACCESS_KEY, UPBIT_SECRET_KEY가 .env에 없습니다.")


def get_upbit_jwt(query_params=None, query_body=None):
    """업비트 API JWT 토큰 생성 (query_hash 포함)
    query_hash는 query string 형식(market=KRW-BTC&side=bid...)이어야 함.
    공식 문서: https://docs.upbit.com/reference/create-authorization-request
    """
    payload = {
        "access_key": UPBIT_ACCESS_KEY,
        "nonce": str(uuid.uuid4()),
    }
    if query_params:
        query_string = urlencode(query_params, doseq=True).replace("%5B%5D=", "[]=")
    elif query_body:
        str_body = {k: str(v) for k, v in query_body.items()}
        query_string = urlencode(str_body, doseq=True).replace("%5B%5D=", "[]=")
    else:
        query_string = ""
    if query_string:
        payload["query_hash"] = hashlib.sha512(query_string.encode("utf-8")).hexdigest()
        payload["query_hash_alg"] = "SHA512"
    return jwt.encode(
        payload, UPBIT_SECRET_KEY, algorithm="HS256"
    )


def get_market_all():
    """마켓 코드 목록 조회 (종목명 매핑용)"""
    url = f"{UPBIT_BASE_URL}/market/all"
    resp = requests.get(url, params={"isDetails": "true"}, timeout=10)
    resp.raise_for_status()
    return resp.json()


def build_name_market_map():
    """종목명/심볼 -> 마켓코드 매핑 생성"""
    markets = get_market_all()
    name_map = {}
    for m in markets:
        mkt = m["market"]
        if not mkt.startswith("KRW-"):
            continue
        korean = m.get("korean_name", "")
        english = m.get("english_name", "")
        symbol = mkt.replace("KRW-", "")
        if korean:
            name_map[korean] = mkt
        if english:
            name_map[english] = mkt
        name_map[symbol] = mkt
        name_map[mkt] = mkt
        name_map[f"{symbol}/KRX"] = mkt
    return name_map


def get_ticker_price(market, retries=3, delay=1):
    """현재가 조회 (인증 불필요)"""
    url = f"{UPBIT_BASE_URL}/ticker"
    for attempt in range(1, retries + 1):
        try:
            resp = requests.get(url, params={"markets": market}, timeout=10)
            if resp.status_code == 200:
                data = resp.json()
                if data:
                    return int(data[0]["trade_price"])
            msg = f"🚨 [현재가 응답 오류] {market} : {resp.status_code} {resp.text}"
            print(msg)
            send_message(msg)
        except Exception as e:
            msg = f"⚠️ [현재가 요청 실패] {market} (시도 {attempt}/{retries}) : {e}"
            print(msg)
            send_message(msg)
        time.sleep(delay)
    msg = f"❌ [현재가 조회 실패] {market} → {retries}회 재시도 후 실패"
    print(msg)
    send_message(msg)
    return None


def get_ticker_prices(markets):
    """여러 마켓 현재가 조회 (markets: ["KRW-ETH", "KRW-ADA", ...])
    일부 마켓이 존재하지 않으면 전체 요청이 404되므로, 개별 조회 후 병합
    업비트 한국: KRW-XXX 형식 (예: KRW-SUI). SUI/KRW → KRW-SUI
    """
    if not markets:
        return {}
    result = {}
    for i, mkt in enumerate(markets):
        if i > 0:
            time.sleep(0.08)
        for attempt in range(2):
            try:
                resp = requests.get(
                    f"{UPBIT_BASE_URL}/ticker",
                    params={"markets": mkt},
                    timeout=8,
                )
                if resp.status_code == 200:
                    data = resp.json()
                    if data:
                        result[mkt] = int(float(data[0]["trade_price"]))
                        break
            except Exception:
                pass
            if attempt == 0:
                time.sleep(0.2)
    return result


def get_accounts():
    """보유 코인/잔고 조회"""
    url = f"{UPBIT_BASE_URL}/accounts"
    token = get_upbit_jwt()
    headers = {"Authorization": f"Bearer {token}"}
    resp = requests.get(url, headers=headers, timeout=10)
    if resp.status_code != 200:
        print(f"⚠️ [계좌 조회 실패] {resp.status_code} {resp.text}")
        return []
    return resp.json()


def get_minute_candles(market, unit=1, count=10):
    """분봉 캔들 조회 (고가/저가 계산용)"""
    url = f"{UPBIT_BASE_URL}/candles/minutes/{unit}"
    params = {"market": market, "count": count}
    resp = requests.get(url, params=params, timeout=10)
    if resp.status_code != 200:
        return None
    data = resp.json()
    if not data:
        return None
    return data


def get_day_candles(market, count=30):
    """일봉 캔들 조회 (기준봉익절 최저가 계산용)"""
    url = f"{UPBIT_BASE_URL}/candles/days"
    params = {"market": market, "count": count}
    resp = requests.get(url, params=params, timeout=10)
    if resp.status_code != 200:
        return None
    data = resp.json()
    if not data:
        return None
    return data


def get_minute_highlow(market, market_name, num_candles=3):
    """최근 분봉 고가/저가 반환"""
    candles = get_minute_candles(market, unit=1, count=max(num_candles, 5))
    if not candles or len(candles) < num_candles:
        return None, None
    recent = candles[:num_candles]
    high = max(float(c["high_price"]) for c in recent)
    low = min(float(c["low_price"]) for c in recent)
    return high, low


def create_order(market, side, ord_type, price=None, volume=None):
    """업비트 공식 API로 주문 생성 (POST /v1/orders)
    https://docs.upbit.com/reference/new-order
    - side: "bid"(매수) | "ask"(매도)
    - ord_type: "price"(시장가 매수) | "market"(시장가 매도) | "limit"(지정가)
    - price: 시장가 매수 시 KRW 금액, 지정가 시 호가
    - volume: 시장가 매도 시 코인 수량, 지정가 시 주문 수량
    """
    body = {"market": market, "side": side, "ord_type": ord_type}
    if price is not None:
        body["price"] = str(int(price))
    if volume is not None:
        body["volume"] = str(volume) if isinstance(volume, float) else str(float(volume))

    token = get_upbit_jwt(query_body=body)
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json; charset=utf-8"}
    url = f"{UPBIT_BASE_URL}/orders"
    resp = requests.post(url, headers=headers, json=body, timeout=10)
    if resp.status_code != 200:
        print(f"🚨 [주문 API 오류] {market} {side} : {resp.status_code} {resp.text}")
        return None
    return resp.json()


def buy_order(market, price_type, quantity=None, price=None):
    """매수 주문 (업비트 공식 API)
    - 시장가(price): price에 KRW 금액 전달
    - 지정가(limit): price, quantity 전달
    """
    try:
        if price_type == "market":
            return create_order(market, "bid", "price", price=int(price))
        return create_order(market, "bid", "limit", price=int(price), volume=float(quantity))
    except Exception as e:
        print(f"🚨 [매수 주문 실패] {market} : {e}")
        return None


def sell_order(market, price_type, quantity, price=None):
    """매도 주문 (업비트 공식 API)
    - 시장가(price): quantity 전달 (코인 수량)
    - 지정가(limit): price, quantity 전달
    """
    try:
        if price_type == "market":
            return create_order(market, "ask", "market", volume=float(quantity))
        return create_order(market, "ask", "limit", price=int(price), volume=float(quantity))
    except Exception as e:
        print(f"🚨 [매도 주문 실패] {market} : {e}")
        return None


def load_excel_with_format(file_path):
    """openpyxl 기반 엑셀 로드 (셀 서식 포함)"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2):
        row_dict = {}
        for idx, cell in enumerate(row):
            col_name = header[idx]
            row_dict[col_name] = cell.value
            if col_name == "감시가격":
                row_dict["감시가격_format"] = cell.number_format
            if col_name == "매매수량":
                row_dict["매매수량_format"] = cell.number_format
            if col_name == "매매가격":
                row_dict["매매가격_format"] = cell.number_format
        rows.append(row_dict)
    return rows


def get_korean_weekday(date_or_str):
    """한글 요일 변환"""
    try:
        if isinstance(date_or_str, str):
            date_only = re.sub(r"\s*\([^)]*\)", "", date_or_str).strip()
            date_obj = pd.to_datetime(date_only)
        elif isinstance(date_or_str, (datetime, pd.Timestamp)):
            date_obj = pd.to_datetime(date_or_str)
        else:
            return ""
        weekday_kr = ["월", "화", "수", "목", "금", "토", "일"]
        return weekday_kr[date_obj.weekday()]
    except Exception:
        return ""


def get_target_price_percent(market, market_name, reason, percent_raw, percent_format, accounts):
    """백분율(수익/손실) 기준 목표가 계산"""
    try:
        percent = float(str(percent_raw).replace("%", "").strip())
        if "%" in str(percent_format):
            percent *= 100
        currency = market.replace("KRW-", "")
        matched = next((a for a in accounts if a.get("currency") == currency), None)
        if not matched:
            raise ValueError("보유 종목이 아님 (매수가 기준 계산 불가)")
        buy_price = float(matched.get("avg_buy_price", 0))
        target_price = int(buy_price * (1 + percent / 100))
        print(f"  📈 [매매가격 계산] [*{market_name}*] {reason} : 매수가 {buy_price:,.0f}원, 감시가격 {percent:+.1f}% → 목표가 {target_price:,}원")
        return target_price
    except Exception as e:
        msg = f"🚨 [*{market_name}*] {reason} : 백분율 감시가격 계산 실패 → '{percent_raw}' ➡️ {e}"
        print(msg)
        send_message(msg)
        return None


def get_target_price_sctp(market, market_name, reason, target_price_raw, from_date_raw):
    """기준봉익절: 기준일 이후 최저가 + 한콤마"""
    try:
        try:
            base_date_obj = pd.to_datetime(re.sub(r"\s*\([^)]*\)", "", str(from_date_raw)).strip())
            base_date = base_date_obj.strftime("%Y-%m-%d")
            base_ts = pd.Timestamp(base_date)
        except Exception as e:
            raise ValueError(f"기준일자 파싱 실패: '{from_date_raw}' ➡️ {e}")
        try:
            price_add = int(float(target_price_raw))
        except ValueError:
            raise ValueError(f"감시가격(한콤마)이 숫자가 아님: '{target_price_raw}'")
        candles = get_day_candles(market, count=100)
        if not candles:
            raise ValueError("일봉 데이터 조회 실패")
        lows = []
        for c in candles:
            c_time = c.get("candle_date_time_kst", c.get("candle_date_kst", ""))
            c_ts = pd.Timestamp(c_time[:10]) if c_time else None
            if c_ts is not None and c_ts >= base_ts:
                lows.append(float(c["low_price"]))
        if not lows:
            raise ValueError("기준일 이후 데이터 없음")
        lowest = int(min(lows))
        target_price = lowest + price_add
        print(f"📌 [기준봉 익절가 계산] [*{market_name}*] {reason} ➡️ 목표가 {target_price:,} (최저가 {lowest:,} + 한콤마 {price_add:,})")
        return target_price
    except Exception as e:
        msg = f"🚨 [기준봉 익절가 계산 실패] [*{market_name}*] : {e}"
        print(msg)
        send_message(msg)
        return None


def process_sctp_trade_type(row, market, market_name, reason, target_price_raw_format):
    """기준봉익절 매매구분 처리"""
    trade_type_str = "매도(기준봉익절)"
    target_price_raw = row["감시가격"]
    from_date_raw = row["감시조건"]
    if not from_date_raw or str(from_date_raw).strip() in ["", "None", "NaT"]:
        weekday_kor = get_korean_weekday(datetime.today())
        from_date_raw = datetime.today().strftime(f"%Y-%m-%d ({weekday_kor})")
        row["감시조건"] = from_date_raw
    target_price = get_target_price_sctp(market, market_name, reason, target_price_raw, from_date_raw)
    row["감시조건"] = "이상"
    return target_price, trade_type_str


def parse_buy_quantity(market_name, reason, trade_qty, trade_unit):
    """매수 수량 파싱 (매매단위: 개=갯수, KRW=원화, %=KRW잔고 백분율)
    반환: {"unit": "개"|"KRW"|"%", "value": float} 또는 None
    """
    try:
        unit = str(trade_unit or "").strip().upper()
        if unit not in ("개", "KRW", "%"):
            unit = "개"
        try:
            val = float(str(trade_qty).replace(",", "").strip())
        except (ValueError, TypeError):
            msg = f"🚨 [*{market_name}*] {reason} : 매매수량 숫자 오류 → '{trade_qty}'"
            print(msg)
            send_message(msg)
            return None
        if val <= 0:
            msg = f"🚨 [*{market_name}*] {reason} : 매매수량이 0 이하"
            print(msg)
            send_message(msg)
            return None
        if unit == "%":
            if 0 < val <= 1.0:
                val = val * 100
            if val <= 0 or val > 100:
                msg = f"🚨 [*{market_name}*] {reason} : 매매단위 % 는 1~100 범위"
                print(msg)
                send_message(msg)
                return None
        return {"unit": unit, "value": val}
    except Exception as e:
        msg = f"🚨 [*{market_name}*] {reason} : 매매수량 해석 실패: '{trade_qty}' → {e}"
        print(msg)
        send_message(msg)
        return None


def parse_sell_quantity(market_name, reason, trade_qty, trade_unit, held_qty, market_price):
    """매도 수량 파싱 (매매단위: 개=갯수, KRW=원화, %=보유수량 백분율)
    반환: 매도할 코인 수량 (float) 또는 0(에러/스킵)
    """
    try:
        unit = str(trade_unit or "").strip().upper()
        if unit not in ("개", "KRW", "%"):
            unit = "개"
        try:
            val = float(str(trade_qty).replace(",", "").replace("%", "").strip())
        except (ValueError, TypeError):
            msg = f"🚨 [*{market_name}*] {reason} : 매매수량 숫자 오류 → '{trade_qty}'"
            print(msg)
            send_message(msg)
            return 0
        if val <= 0:
            msg = f"🚨 [*{market_name}*] {reason} : 매도 수량이 0"
            print(msg)
            send_message(msg)
            return 0

        if unit == "%":
            if 0 < val <= 1.0:
                val = val * 100
            if val <= 0 or val > 100:
                msg = f"🚨 [*{market_name}*] {reason} : 매매단위 % 는 1~100 범위"
                print(msg)
                send_message(msg)
                return 0
            if val >= 99.999999:
                sell_qty = held_qty
            else:
                sell_qty = held_qty * (val / 100)
                sell_qty = round(sell_qty, 8)
        elif unit == "KRW":
            if market_price <= 0:
                msg = f"🚨 [*{market_name}*] {reason} : 시세 조회 실패 (KRW 단위 매도 불가)"
                print(msg)
                send_message(msg)
                return 0
            sell_qty = val / market_price
            sell_qty = round(sell_qty, 8)
        else:
            sell_qty = round(val, 8)

        if sell_qty <= 0:
            msg = f"🚨 [*{market_name}*] {reason} : 매도 수량이 0"
            print(msg)
            send_message(msg)
            return 0
        if unit == "개" and sell_qty > held_qty:
            msg = (
                f"⚠️ [*{market_name}*] {reason} : 매도 수량 초과 → 주문 스킵\n"
                f"        보유수량: {held_qty:.8f} | 매도 요청: {sell_qty:.8f}"
            )
            print(msg)
            send_message(msg)
            return 0
        if sell_qty > held_qty:
            sell_qty = held_qty
        return round(sell_qty, 8) if sell_qty < held_qty else held_qty
    except Exception as e:
        msg = f"🚨 [*{market_name}*] {reason} : 매매수량 해석 실패: '{trade_qty}' → {e}"
        print(msg)
        send_message(msg)
        return 0


def format_duration(seconds):
    """초 → 일/시간/분 문자열"""
    days = seconds // (24 * 3600)
    remainder = seconds % (24 * 3600)
    hours = remainder // 3600
    minutes = (remainder % 3600) // 60
    parts = []
    if days > 0:
        parts.append(f"{int(days)}일")
    if hours > 0:
        parts.append(f"{int(hours)}시간")
    if minutes > 0:
        parts.append(f"{int(minutes)}분")
    return " ".join(parts) or "0분"


def format_result_dict(obj):
    return json.dumps(obj, indent=2, ensure_ascii=False)


def format_holdings_message(accounts, market=None):
    """보유자산 메시지 포맷 (업비트 화면 참고, 테이블 형태)
    market 지정 시 해당 코인+KRW만, None이면 전체
    """
    rows_data = []

    for a in accounts:
        currency = a.get("currency", "")
        bal = float(a.get("balance", 0)) + float(a.get("locked", 0))
        avg = float(a.get("avg_buy_price", 0))
        if currency == "KRW":
            continue
        elif bal > 0 and (not market or market == f"KRW-{currency}"):
            buy_amt = bal * avg
            rows_data.append((currency, bal, avg, buy_amt, 0, 0, f"KRW-{currency}"))

    if market and market.startswith("KRW-"):
        coin = market.replace("KRW-", "")
        if not any(r[0] == coin for r in rows_data):
            held = next((a for a in accounts if a.get("currency") == coin), None)
            if held:
                bal = float(held.get("balance", 0)) + float(held.get("locked", 0))
                avg = float(held.get("avg_buy_price", 0))
                rows_data.append((coin, bal, avg, bal * avg, 0, 0, market))

    if not rows_data:
        return "        (보유 없음)"

    markets_to_fetch = [r[6] for r in rows_data if r[6]]
    prices = get_ticker_prices(markets_to_fetch) if markets_to_fetch else {}

    header = "| 보유자산 |    보유수량   |   평가금액   |      평가손익      |"
    sep = "|----------|---------------|--------------|--------------------|"
    lines = [header, sep]

    out_rows = []
    for row in rows_data:
        cur, bal, avg, buy_amt, _, _, mkt = row
        price = prices.get(mkt, 0) if mkt else 0
        val_amt = bal * price if price else 0
        if buy_amt > 0 and price:
            pl = (val_amt - buy_amt) / buy_amt * 100
            sign = "+" if pl >= 0 else ""
            pl_pct = f"{sign}{pl:.2f}%"
        else:
            pl_pct = "-"

        qty_str = f"{bal:.8f}".rstrip("0").rstrip(".")
        val_str = f"{val_amt:,.0f}원"
        out_rows.append((cur, qty_str, val_str, pl_pct, val_amt))

    out_rows.sort(key=lambda x: x[4], reverse=True)
    for cur, qty_str, val_str, pl_pct, _ in out_rows:
        pl_display = pl_pct if isinstance(pl_pct, str) else str(pl_pct)
        lines.append(f"| {cur:6} | {qty_str:>14} | {val_str:>12} | {pl_display:>18} |")

    return "\n".join(lines)


atexit.register(lambda: send_message(f"🔴 [*{SCRIPT_NAME}*] 스크립트 정상 종료 ({datetime.now().strftime('%Y-%m-%d %H:%M:%S')}) {get_runtime_info()}"))
signal.signal(signal.SIGINT, lambda s, f: (send_message(f"🔴 [*{SCRIPT_NAME}*] 스크립트 종료 (Ctrl+C) {get_runtime_info()}"), sys.exit(0)))
signal.signal(signal.SIGTERM, lambda s, f: (send_message(f"🔴 [*{SCRIPT_NAME}*] 스크립트 종료 (SIGTERM) {get_runtime_info()}"), sys.exit(0)))


def main():
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"임계값 엑셀 파일 없음: {EXCEL_PATH}")

    all_rows = load_excel_with_format(EXCEL_PATH)
    rows = all_rows
    total_count = len(all_rows)
    watch_count = sum(1 for r in rows if str(r.get("감시중", "")).strip().upper() == "O")

    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    text = (
        f"✨ [*{SCRIPT_NAME}*] 업비트 감시 주문 스크립트 ✨\n"
        f"          💾 스크립트 시작 ({now})\n"
        f"           🖥️ {get_runtime_info()}\n"
        f"           🟢 [감시 대상] 등록 {total_count:,}건 ➡️ *감시중 {watch_count:,}건* ({now}) 🟢"
    )
    send_message(text)

    accounts = get_accounts()
    msg_holdings_start = f"📊 [보유잔고] 스크립트 시작 시:\n{format_holdings_message(accounts)}"
    print(msg_holdings_start)
    send_message(msg_holdings_start)

    name_market_map = build_name_market_map()
    sent_first = False
    last_status_hour = None

    while True:
        accounts = get_accounts()
        krw_balance = sum(float(a.get("balance", 0)) + float(a.get("locked", 0)) for a in accounts if a["currency"] == "KRW")
        print(f"🏚️ [계좌 잔고] KRW: {krw_balance:,.0f}원")

        print("\n🗂️ [보유 코인]")
        for a in accounts:
            if a["currency"] == "KRW":
                continue
            bal = float(a.get("balance", 0)) + float(a.get("locked", 0))
            avg = float(a.get("avg_buy_price", 0))
            if bal > 0:
                print(f"    💵 {a['currency']} : 보유 {bal:.8f} / 매수가 {avg:,.0f}원")

        for row in rows:
            if str(row.get("감시중", "")).strip().upper() != "O":
                continue

            stock_name = str(row.get("종목명", "")).strip()
            reason = str(row.get("감시사유", "")).strip()
            trade_type = str(row.get("매매구분", "")).strip()
            trade_type_str = trade_type
            target_price_raw = str(row.get("감시가격", "")).strip()
            target_price_raw_format = row.get("감시가격_format", "")
            valid_until = row.get("유효기간")

            today = datetime.today().date()
            if pd.isnull(valid_until) or str(valid_until).strip() == "":
                print(f"⚠️ [유효기간 없음] 제외: {stock_name} ({reason})")
                continue
            try:
                expiry = pd.to_datetime(valid_until).date()
            except Exception as e:
                print(f"❌ [유효기간 파싱 오류] {stock_name} ({reason}) → {e}")
                continue
            if expiry < today:
                print(f"⏳ [유효기간 경과] 제외: {stock_name} ({reason})")
                continue

            market = name_market_map.get(stock_name)
            if not market:
                for k, v in name_market_map.items():
                    if k.upper() == stock_name.upper():
                        market = v
                        break
            if not market:
                msg = f"🚨 [{stock_name}] {reason} : 마켓코드 매핑 실패"
                print(msg)
                send_message(msg)
                row["감시중"] = "X"
                continue

            target_price = None

            if trade_type == "기준봉익절":
                target_price, trade_type_str = process_sctp_trade_type(row, market, stock_name, reason, target_price_raw_format)
            elif "%" in str(target_price_raw_format) or "%" in str(target_price_raw):
                target_price = get_target_price_percent(market, stock_name, reason, target_price_raw, target_price_raw_format, accounts)
            elif any(x in str(target_price_raw_format) for x in ["원", "₩", "#"]) or str(target_price_raw).replace(".", "", 1).replace("-", "", 1).isdigit():
                try:
                    target_price = int(float(str(target_price_raw).replace("원", "").replace(",", "").strip()))
                except Exception as e:
                    msg = f"🚨 [*{stock_name}*] {reason} : 감시가격(원화) 해석 실패 → {target_price_raw}"
                    print(msg)
                    send_message(msg)
                    row["감시중"] = "X"
                    continue

            if target_price is None:
                msg = f"🚨 [*{stock_name}*] {reason} : target_price is None"
                print(msg)
                send_message(msg)
                row["감시중"] = "X"
                continue

            market_price = get_ticker_price(market)
            if market_price is None:
                continue

            condition = str(row.get("감시조건", "")).strip()
            recent_high, recent_low = get_minute_highlow(market, stock_name, NUM_CANDLES_PRICE)

            if recent_high is not None and recent_low is not None:
                if condition == "이상":
                    condition_met = recent_high >= target_price
                    msg_highlow = f"최근 {NUM_CANDLES_PRICE}분 고가 {recent_high:,.0f}원"
                elif condition == "이하":
                    condition_met = recent_low <= target_price
                    msg_highlow = f"최근 {NUM_CANDLES_PRICE}분 저가 {recent_low:,.0f}원"
                else:
                    print(f"🚨 [감시조건 오류] {stock_name} - '{condition}'")
                    continue
            else:
                if condition == "이상":
                    condition_met = market_price >= target_price
                elif condition == "이하":
                    condition_met = market_price <= target_price
                else:
                    continue
                msg_highlow = f"현재가 {market_price:,.0f}원 (fallback)"

            if not condition_met:
                continue

            now = datetime.now()
            target_price_fmt = f"{target_price:,.0f}원"
            current_price_fmt = f"{market_price:,.0f}원"
            reason_text = f" {reason}" if reason else ""
            stock_url = f"https://upbit.com/exchange?code=CRIX.UPBIT.{market}"
            stock_name_link = f"<{stock_url}|{stock_name}>"

            trade_qty_raw = row.get("매매수량", "")
            trade_unit = str(row.get("매매단위", "") or "").strip().upper()
            if not trade_unit:
                fmt = str(row.get("매매수량_format", "") or "")
                if "%" in fmt:
                    trade_unit = "%"
                elif "KRW" in fmt or "원" in fmt:
                    trade_unit = "KRW"
                else:
                    trade_unit = "개"
            if trade_unit not in ("개", "KRW", "%"):
                trade_unit = "개"
            try:
                qty_val = float(str(trade_qty_raw).replace(",", "").replace("%", "").strip())
            except (ValueError, TypeError):
                qty_val = str(trade_qty_raw)
            if isinstance(qty_val, float):
                if trade_unit == "%":
                    trade_qty_display = f"{qty_val:.0f}%"
                elif trade_unit == "KRW":
                    trade_qty_display = f"{qty_val:,.0f}원"
                else:
                    trade_qty_display = f"{qty_val}개"
            else:
                trade_qty_display = str(qty_val)

            valid_until_str = pd.to_datetime(valid_until).strftime("%Y-%m-%d") if pd.notnull(valid_until) else "N/A"

            msg = (
                f"  🔍 [*매매조건 감지*] {stock_name_link} - 감시사유: {reason_text} ({now.strftime('%m-%d %H:%M:%S')})\n"
                f"                                매매구분: *{trade_type_str}*   감시가격: *{target_price_fmt} {condition}*\n"
                f"                                현재가: *{current_price_fmt}* [{msg_highlow}]\n"
                f"                                매매수량: {trade_qty_display} | 유효기간: {valid_until_str}"
            )
            print(msg)
            send_message(msg)

            if trade_type == "매수":
                buy_info = parse_buy_quantity(stock_name, reason, trade_qty_raw, trade_unit)
                if not buy_info:
                    row["감시중"] = "X"
                    continue
                order_price = str(row.get("매매가격", "")).strip().lower()
                price_type = "market" if order_price == "market" else "limit"
                unit = buy_info["unit"]
                val = buy_info["value"]
                if price_type == "market":
                    if unit == "개":
                        krw_amt = int(val * market_price)
                    elif unit == "KRW":
                        krw_amt = int(val)
                    else:
                        krw_amt = int(krw_balance * val / 100)
                    if krw_amt < UPBIT_MIN_ORDER_KRW:
                        msg = (
                            f"⚠️ [*{stock_name}*] {reason} : 주문금액 미달 → 주문 스킵\n"
                            f"        매매수량: {trade_qty_display} → 주문금액 약 {krw_amt:,}원 "
                            f"(업비트 최소 주문금액 {UPBIT_MIN_ORDER_KRW:,}원)"
                        )
                        print(msg)
                        send_message(msg)
                        row["감시중"] = "X"
                        continue
                    result = buy_order(market, "market", price=krw_amt)
                else:
                    try:
                        price_val = int(order_price)
                    except ValueError:
                        msg = f"🚨 [{stock_name}] 지정가 변환 실패: '{order_price}'"
                        print(msg)
                        send_message(msg)
                        row["감시중"] = "X"
                        continue
                    if unit == "개":
                        order_qty = val
                    elif unit == "KRW":
                        order_qty = val / price_val
                    else:
                        order_qty = (krw_balance * val / 100) / price_val
                    order_amt = int(order_qty * price_val)
                    if order_amt < UPBIT_MIN_ORDER_KRW:
                        msg = (
                            f"⚠️ [*{stock_name}*] {reason} : 주문금액 미달 → 주문 스킵\n"
                            f"        매매수량: {trade_qty_display} × 가격 {price_val:,}원 = 약 {order_amt:,}원 "
                            f"(업비트 최소 주문금액 {UPBIT_MIN_ORDER_KRW:,}원)"
                        )
                        print(msg)
                        send_message(msg)
                        row["감시중"] = "X"
                        continue
                    result = buy_order(market, "limit", quantity=order_qty, price=price_val)

            elif trade_type in ["매도", "기준봉익절"]:
                currency = market.replace("KRW-", "")
                held = next((a for a in accounts if a["currency"] == currency), None)
                held_qty = float(held.get("balance", 0)) + float(held.get("locked", 0)) if held else 0
                if held_qty <= 0:
                    msg = f"⚠️ [*{stock_name}*] {reason} : 보유수량 0 → 주문 제외"
                    print(msg)
                    send_message(msg)
                    row["감시중"] = "X"
                    continue
                order_qty = parse_sell_quantity(stock_name, reason, trade_qty_raw, trade_unit, held_qty, market_price)
                if not order_qty:
                    row["감시중"] = "X"
                    continue
                order_price = str(row.get("매매가격", "")).strip().lower()
                price_type = "market" if order_price == "market" else "limit"
                if price_type == "market":
                    order_amt = int(order_qty * market_price)
                else:
                    try:
                        price_val = int(order_price)
                    except ValueError:
                        msg = f"🚨 [{stock_name}] 지정가 변환 실패: '{order_price}'"
                        print(msg)
                        send_message(msg)
                        row["감시중"] = "X"
                        continue
                    order_amt = int(order_qty * price_val)
                if order_amt < UPBIT_MIN_ORDER_KRW:
                    msg = (
                        f"⚠️ [*{stock_name}*] {reason} : 주문금액 미달 → 주문 스킵\n"
                        f"        매매수량: {trade_qty_display} → 주문금액 약 {order_amt:,}원 "
                        f"(업비트 최소 주문금액 {UPBIT_MIN_ORDER_KRW:,}원)"
                    )
                    print(msg)
                    send_message(msg)
                    row["감시중"] = "X"
                    continue
                if price_type == "market":
                    result = sell_order(market, "market", order_qty)
                else:
                    result = sell_order(market, "limit", order_qty, price_val)
            else:
                msg = f"🚨 [*{stock_name}*] 알 수 없는 매매구분: '{trade_type}'"
                print(msg)
                send_message(msg)
                continue

            row["감시중"] = "X"

            if result is None:
                msg_after = f"🔴 [주문 결과] [*{stock_name}*] {reason} : 주문 실패"
                print(msg_after)
                send_message(msg_after)
            else:
                msg_after = f"🟢 [주문 결과] [*{stock_name}*] {reason} :\n{format_result_dict(result)}"
                print(msg_after)
                send_message(msg_after)
                accounts_after = get_accounts()
                msg_holdings = (
                    f"📊 [보유수량] [*{stock_name}*] 주문 후:\n{format_holdings_message(accounts_after, market)}"
                )
                print(msg_holdings)
                send_message(msg_holdings)

        if not sent_first:
            send_message(f"🟡 {datetime.now().strftime('%m-%d %H:%M:%S')} - 최초 감시 완료 ⏱️")
            sent_first = True

        now = datetime.now()
        if SEND_HOURLY_MSG and last_status_hour != now.hour:
            watch_count = sum(1 for r in rows if str(r.get("감시중", "")).strip().upper() == "O")
            send_message(f"✨ [*{SCRIPT_NAME}*] [정시 알림] 감시중 {watch_count:,}건 ({now.strftime('%Y-%m-%d %H:%M:%S')})")
            last_status_hour = now.hour

        print(f"🟡 {now.strftime('%m-%d %H:%M:%S')} - 감시 완료. 1분 대기 후 계속...")
        time.sleep(60)


if __name__ == "__main__":
    main()
